# import des packages
from openpyxl import *
from tkinter import *



# Ouverture du fichier existant
wb = load_workbook('dd.xlsx')


sheet = wb.active


def excel():
	
	#Dimensions des colonnes dans le tableur
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# Header dans le tableur
	sheet.cell(row=1, column=1).value = "Nom"
	sheet.cell(row=1, column=2).value = "Prenom"
	sheet.cell(row=1, column=3).value = "Classe"
	sheet.cell(row=1, column=4).value = "Numero"


# Fonction focus (curseur)
def focus1(event):
	course_field.focus_set()


# Fonction focus (curseur)
def focus2(event):
	sem_field.focus_set()


# Fonction focus (curseur)
def focus3(event):
	form_no_field.focus_set()


# Fonction focus (curseur)
def focus4(event):
	contact_no_field.focus_set()


# Fonction focus (curseur)
def focus5(event):
	email_id_field.focus_set()


# Fonction focus (curseur)
def focus6(event):
	address_field.focus_set()


# Suppression des contenus après ajout
def clear():
	
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)


# Ecriture des données sur le tableur
def insert():
	
	# si input vide
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:

		
		current_row = sheet.max_row
		current_column = sheet.max_column

		
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()

		# Enregistrement du fichier
		wb.save('dd.xlsx')

		name_field.focus_set()

		clear()


# Fenêtre
if __name__ == "__main__":
	
	
	root = Tk()

	# Font
	root.configure(background='light green')

	# Titre
	root.title("registration form")

	# Dimensions
	root.geometry("500x300")

	excel()

	# Text à afficher sur la fenêtre
	heading = Label(root, text="Form", bg="light green")

	# Text à afficher sur le nom
	name = Label(root, text="Nom", bg="light green")

	# Text à afficher sur le prenom
	course = Label(root, text="Prenom", bg="light green")

	# Text à afficher sur la classe
	sem = Label(root, text="Classe", bg="light green")

	# Text à afficher sur le numero
	form_no = Label(root, text="Numero", bg="light green")

	

	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)

	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)

	
	name_field.bind("<Return>", focus1)


	course_field.bind("<Return>", focus2)

	sem_field.bind("<Return>", focus3)


	form_no_field.bind("<Return>", focus4)


	contact_no_field.bind("<Return>", focus5)

	email_id_field.bind("<Return>", focus6)


	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")

	excel()

	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=8, column=1)


	root.mainloop()
